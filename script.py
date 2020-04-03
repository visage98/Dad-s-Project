import csv
import sys
import openpyxl
import re

from lcoConfig import validity_month, input_file, output_file, output_column1, output_column2, can_column, stb_column, max_rows
from bot import LcoBot

def getDict():
    can_dict = {}
    stb_dict = {}

    wb = openpyxl.load_workbook(output_file)
    sheet = wb.active
    pattern1 = r'CAN\d+'
    pattern2 = r'STB\d+'
    index = 1
    for c1,c2 in sheet.iter_rows(min_row=1, min_col=1, max_row = max_rows, max_col = 2):
        s1 = str(c1.value).strip().upper()
        s2 = 'STB'+str(c2.value).strip()
        if re.search(pattern1, s1):
            can_dict[s1] = index
        if re.search(pattern2, s2):
            stb_dict[s2] = index
        index=index+1

    return can_dict, stb_dict
        

can_dict, stb_dict = getDict()

bot = LcoBot(validity_month)

file = open(input_file)
errorFile = open('logFile.csv','w')
csvreader = csv.reader(file)
csvwriter = csv.writer(errorFile)

wb = openpyxl.load_workbook(output_file)
sheet = wb.active

count=0
csvwriter.writerow(["Account Number"])
next(csvreader)
for row in csvreader:   
    account = row[0].strip().upper()
    print(count, account, end=" : ")
    if account in can_dict or account in stb_dict:
        if account[:3]=='CAN':
            cell_loc1 = output_column1+str(can_dict[account])
            cell_loc2 = output_column2+str(can_dict[account])
        else:
            cell_loc1 = output_column1+str(stb_dict[account])
            cell_loc2 = output_column2+str(stb_dict[account])

        if not sheet[cell_loc1].value:
            period = bot.checkBox(account)
            if len(period)==0:
                print("ERROR : LCO")
                csvwriter.writerow([account])
            else:
                split_period = re.search(r"(..)\/(..)\/(....) TO (..)\/(..)\/(....)", period)
                period = "{}-{}-{} TO {}-{}-{}".format(split_period[1], split_period[2], split_period[3][2:], split_period[4], split_period[5], split_period[6][2:])
                print("WRITTEN", cell_loc1, period)
                sheet[cell_loc1] = period
                sheet[cell_loc2] = "BAL"
        else:
            print("Already Filled")
    else:
        print("ERROR : EXCEL")
        csvwriter.writerow([account])

    count=count+1
    # if(count==10):
    #     break

wb.save(output_file)
errorFile.close()
file.close()